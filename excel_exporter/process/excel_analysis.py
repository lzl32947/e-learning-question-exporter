import os
import traceback
from enum import Enum
from typing import List, Optional

import openpyxl

from .types import HeaderType
from ..errors import ProcessError, InputHeaderError
from ..question import Question, Question_AnQuanZhunRu, Question_AnGui


class OutputType:
    def __init__(self, title: str, store_path: str, question: List[Question]):
        self.title = title
        self.store_path = store_path
        self.question = question


class ExcelAnalysis:
    def __init__(self):
        raise RuntimeError("Error >> This class is not allowed to be instantiated!")

    @staticmethod
    def read_in_questions(workbook: openpyxl.workbook.workbook.Workbook, sheet_index: int, input_type: HeaderType, with_header: bool) -> List[Question]:
        question_list = []

        sheet = workbook.worksheets[sheet_index]
        if with_header:
            sheet_iter = sheet.iter_rows(min_row=2, values_only=True)
        else:
            sheet_iter = sheet.iter_rows(min_row=1, values_only=True)

        if input_type == HeaderType.AnQuanZhunRu:
            for row in sheet_iter:
                try:
                    index = int(row[0])
                except ValueError:
                    index = -1

                question = row[1]
                company = row[2]
                question_type = row[3]
                easiness = row[4]
                options = row[5]
                answer = row[6]
                source = row[7]

                question_list.append(
                    Question_AnQuanZhunRu(index, question, company, question_type, easiness, options, answer, source)
                )
        elif input_type == HeaderType.AnGui:
            for row in sheet_iter:
                try:
                    index = int(row[0])
                except ValueError:
                    index = -1

                l1 = row[1]
                l2 = row[2]
                question_catalog = row[3]
                question_type = row[4]
                question = row[5]
                options = row[6]
                answer = row[7]
                source = row[8]

                question_list.append(
                    Question_AnGui(index, l1, l2, question_catalog, question_type, question, options, answer, source)
                )
        else:
            raise InputHeaderError(f"The input header type is not supported! Found {input_type}!")
        return question_list

    @staticmethod
    def analysis(excel_path: str, header_type: HeaderType, with_header: bool) -> Optional[OutputType]:
        try:
            question_list = []
            book = openpyxl.load_workbook(excel_path)
            sheets = len(book.worksheets)
            for index in range(sheets):
                questions = ExcelAnalysis.read_in_questions(book, index, header_type, with_header)
                question_list.extend(questions)
            return OutputType(os.path.basename(excel_path), excel_path, question_list)

        except (FileNotFoundError, Exception) as e:

            raise ProcessError("Error found in Process: " + traceback.format_exc())
